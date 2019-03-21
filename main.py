# import PyQt5
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, \
    QVBoxLayout, QPushButton, QFileDialog, QMessageBox, \
    QTabWidget, QLabel, QGridLayout, QComboBox, QSpinBox, \
    QLineEdit, QDialogButtonBox, QDataWidgetMapper
from PyQt5.QtCore import Qt
from PyQt5.QtSql import QSqlQuery, QSqlError, QSqlRelationalDelegate
from shutil import copyfile
from openpyxl import load_workbook, Workbook

from SQLView import SQLView
from Models import ScheduleModel, PulpitsModel, GroupsModel
from Database import createConnection, deleteAllTables, createTables
import Database
from FilterDlg import FilterDlg


#    schedule
# id, id_pulpit, id_group, id_week, id_day, n_lesson, lesson
# id
# id_pulpit
# id_group
# id_week
# id_day
# n_lesson
# lesson
#    pulpits   groups
# id
# name
#    weeks   days
# id
# name


class SQLWindow(QMainWindow):
    currDBName = ''
    db = None

    scheduleView = None
    sMapper = None
    sMappingPairs = []

    pulpitsView = None
    pMapper = None
    pMappingPairs = []

    groupsView = None
    gMapper = None
    gMappingPairs = []

    filterDlg = None
    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)

        self.setWindowTitle('SQL')
        self.setWidgets()
        self.setMenu()
        self.resize(520, 400)

    def setWidgets(self):
        '''Расстановка виджетов'''

        self.filterDlg = FilterDlg(self)

        # w = QWidget()
        # vLay = QVBoxLayout()

        # таб бар с 3-мя табл: расписание, кафедры, группы
        tabWidget = QTabWidget()


        # 1-й виджет с расписанием и управлением табл
        wSchedule = QWidget()
        vLaySchedule = QVBoxLayout()
        # таблица
        self.scheduleView = SQLView()
        self.scheduleView.setModel(ScheduleModel())
        vLaySchedule.addWidget(self.scheduleView)
        # управление
        gLaySchedule = QGridLayout()
        # добавление надписей
        gLaySchedule.addWidget(QLabel('Кафедра:'), 0, 0, Qt.AlignRight)
        gLaySchedule.addWidget(QLabel('Группа:'), 0, 2, Qt.AlignRight)
        gLaySchedule.addWidget(QLabel('Неделя:'), 1, 0, Qt.AlignRight)
        gLaySchedule.addWidget(QLabel('День:'), 1, 2, Qt.AlignRight)
        gLaySchedule.addWidget(QLabel('№ пары:'), 2, 0, Qt.AlignRight)
        gLaySchedule.addWidget(QLabel('Пара:'), 2, 2, Qt.AlignRight)
        gLaySchedule.addWidget(QLabel('id:'), 3, 0, Qt.AlignRight)
        # создание и вставка в layout виджетов настройки
        sId = QSpinBox(); \
                sId.setRange(1, 100000)
        sPulpit = QComboBox()
        sGroup = QComboBox()
        sWeek = QComboBox()
        sDay = QComboBox()
        sNLesson = QSpinBox(); \
                sNLesson.setRange(1, 10)
        sLesson = QLineEdit()
        # вставка в GridLayout
        gLaySchedule.addWidget(sPulpit, 0, 1)
        gLaySchedule.addWidget(sGroup, 0, 3)
        gLaySchedule.addWidget(sWeek, 1, 1)
        gLaySchedule.addWidget(sDay, 1, 3)
        gLaySchedule.addWidget(sNLesson, 2, 1)
        gLaySchedule.addWidget(sLesson, 2, 3)
        gLaySchedule.addWidget(sId, 3, 1)
        # создание и вставка группы кнопок
        sBtns = QDialogButtonBox()
        sbSave = QPushButton('Записать')
        sbRevert = QPushButton('Вернуть')
        sbNewLine = QPushButton('Новая строка')
        sbDelLines = QPushButton('Удалить строки')
        # connects
        sbSave.clicked.connect(self.scheduleView.saveData)
        sbRevert.clicked.connect(self.scheduleView.revertAll)
        sbNewLine.clicked.connect(self.scheduleView.insertRow)
        sbDelLines.clicked.connect(self.scheduleView.deleteRow)
        # добавление кнопок в группу
        sBtns.addButton(sbSave, QDialogButtonBox.ActionRole)
        sBtns.addButton(sbRevert, QDialogButtonBox.ActionRole)
        sBtns.addButton(sbNewLine, QDialogButtonBox.ActionRole)
        sBtns.addButton(sbDelLines, QDialogButtonBox.ActionRole)
        sBtns.setCenterButtons(True)
        gLaySchedule.addWidget(sBtns, 3, 2, 1, 2)
        gLaySchedule.setColumnStretch(1, 1)
        gLaySchedule.setColumnStretch(3, 1)
        vLaySchedule.addLayout(gLaySchedule)
        # создание маппера для настроек
        self.sMappingPairs.append([sId, 0])
        self.sMappingPairs.append([sPulpit, 1])
        self.sMappingPairs.append([sGroup, 2])
        self.sMappingPairs.append([sWeek, 3])
        self.sMappingPairs.append([sDay, 4])
        self.sMappingPairs.append([sNLesson, 5])
        self.sMappingPairs.append([sLesson, 6])
        self.setScheduleMapper()

        self.scheduleView.currentModelChanged.connect(self.setScheduleMapper)

        wSchedule.setLayout(vLaySchedule)


        # 2-й виджет с кафедрами и управлением табл
        wPulpits = QWidget()
        vLayPulpits = QVBoxLayout()
        # таблица
        self.pulpitsView = SQLView()
        self.pulpitsView.setModel(PulpitsModel())
        vLayPulpits.addWidget(self.pulpitsView)
        # управление
        gLayPulpits = QGridLayout()
        # вставка надписи
        gLayPulpits.addWidget(QLabel('id:'), 0, 0, Qt.AlignRight)
        gLayPulpits.addWidget(QLabel('Название кафедры:'), 0, 2, Qt.AlignRight)
        # вставка edit
        pePulpit = QLineEdit()
        psbId = QSpinBox(); \
                psbId.setRange(1, 100000)
        gLayPulpits.addWidget(psbId, 0, 1)
        gLayPulpits.addWidget(pePulpit, 0, 3)
        # создание группы кнопок
        pBtns = QDialogButtonBox()
        pbSave = QPushButton('Записать')
        pbRevert = QPushButton('Вернуть')
        pbNewLine = QPushButton('Новая строка')
        pbDelLines = QPushButton('Удалить строки')
        # connects
        pbSave.clicked.connect(self.pulpitsView.saveData)
        pbRevert.clicked.connect(self.pulpitsView.revertAll)
        pbNewLine.clicked.connect(self.pulpitsView.insertRow)
        pbDelLines.clicked.connect(self.pulpitsView.deleteRow)
        # добавление в группу кнопок
        pBtns.addButton(pbSave, QDialogButtonBox.ActionRole)
        pBtns.addButton(pbRevert, QDialogButtonBox.ActionRole)
        pBtns.addButton(pbNewLine, QDialogButtonBox.ActionRole)
        pBtns.addButton(pbDelLines, QDialogButtonBox.ActionRole)
        pBtns.setCenterButtons(True)

        gLayPulpits.addWidget(pBtns, 3, 0, 1, 4)
        gLayPulpits.setColumnStretch(3, 1)
        vLayPulpits.addLayout(gLayPulpits)
        # создание маппера
        self.pMappingPairs.append([psbId, 0])
        self.pMappingPairs.append([pePulpit, 1])
        self.setPulpitsMapper()

        self.pulpitsView.currentModelChanged.connect(self.setPulpitsMapper)
        self.pulpitsView.changedData.connect(self.refreshScheduleComboBoxes)
        wSchedule.setLayout(vLaySchedule)




        wPulpits.setLayout(vLayPulpits)

        # 3-й виджет с группами и управл табл
        wGroups = QWidget()
        vLayGroups = QVBoxLayout()
        # таблица
        self.groupsView = SQLView()
        self.groupsView.setModel(GroupsModel())
        vLayGroups.addWidget(self.groupsView)

        # управление
        gLayGroups = QGridLayout()
        # вставка надписи
        gLayGroups.addWidget(QLabel('id:'), 0, 0, Qt.AlignRight)
        gLayGroups.addWidget(QLabel('Название группы:'), 0, 2, Qt.AlignRight)
        # вставка edit
        gsbId = QSpinBox(); \
                gsbId.setRange(1, 100000)
        geGroup = QLineEdit()
        gLayGroups.addWidget(gsbId, 0, 1)
        gLayGroups.addWidget(geGroup, 0, 3)
        # создание группы кнопок
        gBtns = QDialogButtonBox()
        gbSave = QPushButton('Записать')
        gbRevert = QPushButton('Вернуть')
        gbNewLine = QPushButton('Новая строка')
        gbDelLines = QPushButton('Удалить строки')
        # connects
        gbSave.clicked.connect(self.groupsView.saveData)
        gbRevert.clicked.connect(self.groupsView.revertAll)
        gbNewLine.clicked.connect(self.groupsView.insertRow)
        gbDelLines.clicked.connect(self.groupsView.deleteRow)
        # добавление в группу кнопок
        gBtns.addButton(gbSave, QDialogButtonBox.ActionRole)
        gBtns.addButton(gbRevert, QDialogButtonBox.ActionRole)
        gBtns.addButton(gbNewLine, QDialogButtonBox.ActionRole)
        gBtns.addButton(gbDelLines, QDialogButtonBox.ActionRole)
        gBtns.setCenterButtons(True)

        gLayGroups.addWidget(gBtns, 1, 0, 1, 4)
        gLayGroups.setColumnStretch(3, 1)
        vLayGroups.addLayout(gLayGroups)
        # создание маппера
        self.gMappingPairs.append([gsbId, 0])
        self.gMappingPairs.append([geGroup, 1])
        self.setGroupsMapper()

        self.groupsView.currentModelChanged.connect(self.setGroupsMapper)
        self.groupsView.changedData.connect(self.refreshScheduleComboBoxes)

        wGroups.setLayout(vLayGroups)


        # добавление 3-х виджетов в таб виджет
        tabWidget.addTab(wSchedule, 'Расписание')
        tabWidget.addTab(wPulpits, 'Кафедры')
        tabWidget.addTab(wGroups, 'Группы')

        self.setCentralWidget(tabWidget)

    def setMenu(self):
        '''Создание меню'''
        menuBar = self.menuBar()
        mFile = menuBar.addMenu('&Файл')
        mFile.addAction('&Создать пустую БД', self.slot_CreateDB)
        mFile.addAction('&Открыть БД', self.slot_OpenDB, 'Ctrl+O')
        mFile.addAction('Сохранить &БД', self.slot_SaveAsDB, 'Ctrl+S')
        mFile.addSeparator()
        mFile.addAction('Создать &таблицы', self.slot_CreateTables)
        mFile.addSeparator()
        mFile.addAction('&Полный экспорт в *.sql', self.slot_FullExportSql)
        mFile.addAction('Экспорт &данных в *.sql', self.slot_DataExportSql)
        mFile.addAction('&Импорт *.sql', self.slot_ImportSql)
        mFile.addSeparator()
        mFile.addAction('&Выход', self.close)

        mExcel = menuBar.addMenu('&Excel')
        mExcel.addAction('Экспортировать по фильтру', self.slot_ExportExcel)
        mExcel.addAction('Импортировать', self.slot_ImportExcel)

        mFilter = menuBar.addMenu('Ф&ильтр')
        mFilter.addAction('&Выставить фильтр', self.slot_SetFilter)
        mFilter.addAction('&Сбросить фильтр', self.slot_SetDefaultFilter)

        mHelp = menuBar.addMenu('&?')
        mHelp.addAction('&Автор', self.slot_Author)
        mHelp.addAction('&О программе', self.slot_AboutProg, 'F1')

    def refreshScheduleComboBoxes(self):
        '''Обновление моделей для ComboBox Расписания'''
        for i in range(1, 5):
            self.sMappingPairs[i][0].model().select()

    def setScheduleMapper(self):
        '''Создание маппера для расписания (состоит из 6 эл-в)'''
        self.sMapper = QDataWidgetMapper()
        self.sMapper.setModel(self.scheduleView.model())

        self.sMapper.setItemDelegate(QSqlRelationalDelegate())

        for i in range(1, 5):
            rel_model = self.scheduleView.model().relationModel(i)
            self.sMappingPairs[i][0].setModel(rel_model)
            self.sMappingPairs[i][0].setModelColumn(1)

        for pair in self.sMappingPairs:
            self.sMapper.addMapping(pair[0], pair[1])
        self.sMapper.toFirst()

        self.scheduleView.selectionModel().currentRowChanged.connect(self.sMapper.setCurrentModelIndex)
        self.scheduleView.setCurrentIndex(self.scheduleView.model().index(0, 0))

    def setPulpitsMapper(self):
        '''Создание маппера для кафедр'''
        self.pMapper = QDataWidgetMapper()
        self.pMapper.setModel(self.pulpitsView.model())

        self.pMapper.setItemDelegate(QSqlRelationalDelegate())
        for pair in self.pMappingPairs:
            self.pMapper.addMapping(pair[0], pair[1])
        self.pMapper.toFirst()

        self.pulpitsView.selectionModel().currentRowChanged.connect(self.pMapper.setCurrentModelIndex)
        self.pulpitsView.setCurrentIndex(self.pulpitsView.model().index(0, 0))

    def setGroupsMapper(self):
        '''Создание маппера для групп'''
        self.gMapper = QDataWidgetMapper()
        self.gMapper.setModel(self.groupsView.model())

        self.gMapper.setItemDelegate(QSqlRelationalDelegate())
        for pair in self.gMappingPairs:
            self.gMapper.addMapping(pair[0], pair[1])
        self.gMapper.toFirst()

        self.groupsView.selectionModel().currentRowChanged.connect(self.gMapper.setCurrentModelIndex)
        self.groupsView.setCurrentIndex(self.groupsView.model().index(0, 0))

    # MENU FILE
    # with DB
    def slot_CreateDB(self):
        '''Создание пустой БД'''
        # получ имя новой ДБ
        dlg = QFileDialog()
        filename_cort = dlg.getSaveFileName(caption='Выберите имя файла базы данных')
        filename = filename_cort[0]
        if len(filename) == 0:
            return
        # созраняем имя БД в пер-ю и создаем ДБ
        self.currDBName = filename
        self.db = createConnection(db_name=filename)
        if self.db == None:
            QMessageBox.information(self, 'Ошибка', 'Ошибка создания базы данных!')
            return
        # удаляем
        deleteAllTables(self.db)

        self.scheduleView.setModel(ScheduleModel())
        self.pulpitsView.setModel(PulpitsModel())
        self.groupsView.setModel(GroupsModel())

    def slot_CreateTables(self):
        '''Слот создания таблиц в БД'''
        if self.db == None or not createTables(self.db):
            QMessageBox.information(self, 'Ошибка', 'Ошибка создания таблиц!')
            return
        else:
            self.scheduleView.setModel(ScheduleModel())
            self.pulpitsView.setModel(PulpitsModel())
            self.groupsView.setModel(GroupsModel())
            QMessageBox.information(self, 'Сообщение', 'Таблицы созданы!')

    def slot_OpenDB(self):
        '''Открытие БД'''
        # получ имени файла
        dlg = QFileDialog()
        filename_cort = dlg.getOpenFileName(None, 'Выберите базу данных')
        filename = filename_cort[0]
        if len(filename) == 0:
            return
        # соединение
        self.currDBName = filename
        self.db = createConnection(db_name=filename)
        if self.db == None:
            QMessageBox.information(self, 'Ошибка', 'Ошибка открытия базы данных!')
            return

        self.scheduleView.setModel(ScheduleModel())
        self.pulpitsView.setModel(PulpitsModel())
        self.groupsView.setModel(GroupsModel())

    def slot_SaveAsDB(self):
        '''Сохранение текущего состоя БД'''
        # проверка текущего имени файла
        if self.currDBName == None or len(self.currDBName) == 0:
            QMessageBox.information(None, 'Ошибка',
                                    'Неверное имя текущей БД!')
            return
        # получ имени файла
        dlg = QFileDialog()
        filename_cort = dlg.getSaveFileName(caption='Выберите имя файла')
        filename = filename_cort[0]
        if len(filename) == 0:
            return
        # копирование
        try:
            copyfile(self.currDBName, filename)
        except IOError:
            QMessageBox.information(None, 'Ошибка', 'Ошибка копирования файла!')
            return
        except Exception:
            print('Unknown error!')
            return
        # подсоединение к новой БД
        self.currDBName = filename
        db = createConnection(db_name=self.currDBName)
        if db == None:
            QMessageBox.information(self, 'Ошибка', 'Ошибка соединения с новой БД!')
            return

        self.scheduleView.setModel(ScheduleModel())
        self.pulpitsView.setModel(PulpitsModel())
        self.groupsView.setModel(GroupsModel())

    def __writeDataInFile(self, f_desc):
        # запись кафедр
        template_pulpits = 'INSERT INTO pulpits (id, name) VALUES ({0}, "{1}");\n'
        sql = QSqlQuery()
        sql.exec('SELECT * FROM pulpits;')
        if sql.lastError().type() != QSqlError.NoError:
            QMessageBox.information(self, 'Ошибка', 'Ошибка запроса получения кафедр из БД!')
            return False
        while sql.next():
            f_desc.write(template_pulpits.format(sql.value(0), sql.value(1)))
        # запись групп
        template_groups = 'INSERT INTO groups (id, name) VALUES ({0}, "{1}");\n'
        sql = QSqlQuery()
        sql.exec('SELECT * FROM groups;')
        if sql.lastError().type() != QSqlError.NoError:
            QMessageBox.information(self, 'Ошибка', 'Ошибка запроса получения групп из БД!')
            return False
        while sql.next():
            f_desc.write(template_groups.format(sql.value(0), sql.value(1)))
        # запись расписания
        template_schedule = 'INSERT INTO schedule (id, id_pulpit, id_group, id_week, id_day, n_lesson, lesson) ' \
                            'VALUES ({0}, {1}, {2}, {3}, {4}, {5}, "{6}");\n'
        sql = QSqlQuery()
        sql.exec('SELECT * FROM schedule;')
        if sql.lastError().type() != QSqlError.NoError:
            QMessageBox.information(self, 'Ошибка', 'Ошибка запроса получения расписания из БД!')
            return False
        while sql.next():
            f_desc.write(template_schedule.format(sql.value(0), sql.value(1), sql.value(2),
                                                  sql.value(3), sql.value(4), sql.value(5),
                                                  sql.value(6)))
        return True

    # with .sql
    def slot_FullExportSql(self):
        '''Экспорт таблиц и данных'''
        if self.db == None:
            QMessageBox.information(self, 'Ошибка', 'Экспорт невозможен, т.к. БД не открыта!')
            return
        # получение имени файла
        dlg = QFileDialog()
        filename_cort = dlg.getSaveFileName(caption='Выберите имя файла', filter='*.sql')
        filename = filename_cort[0]
        if len(filename) == 0:
            return
        # открытие файла и запись
        try:
            fout = open(filename, 'w')
        except Exception:
            print('exception')
        fout.write(Database.create_schedule_query + '\n')
        fout.write(Database.create_pulpits_query + '\n')
        fout.write(Database.create_groups_query + '\n')
        fout.write(Database.create_weeks_query + '\n')
        fout.write(Database.create_days_query + '\n')
        for insert_cmd in Database.insert_weeks_query:
            fout.write(insert_cmd + '\n')
        for insert_cmd in Database.insert_days_query:
            fout.write(insert_cmd + '\n')
        if not self.__writeDataInFile(fout):
            fout.close()
            return
        fout.close()

    def slot_DataExportSql(self):
        '''Экспорт данных расписания, кафедр, групп в *.sql файл'''
        if self.db == None:
            QMessageBox.information(self, 'Ошибка', 'Экспорт невозможен, т.к. БД не открыта!')
            return
        # получение имени файла
        dlg = QFileDialog()
        filename_cort = dlg.getSaveFileName(caption='Выберите имя файла', filter='*.sql')
        filename = filename_cort[0]
        if len(filename) == 0:
            return
        # открытие файла, запись данных
        try:
            fout = open(filename, 'w')
        except Exception:
            QMessageBox.information(self, 'Ошибка', 'Ошибка открытия файла!')
            return
        if not self.__writeDataInFile(fout):
            fout.close()
            return

        fout.close()

    def slot_ImportSql(self):
        '''Импорт (выполнение команд) файла *.sql'''
        if self.db == None:
            QMessageBox.information(self, 'Ошибка', 'Импорт невозможен, т.к. БД не открыта!')
            return
        # получение имени файла
        dlg = QFileDialog()
        filename_cort = dlg.getOpenFileName(caption='Выберите файл', filter='*.sql')
        filename = filename_cort[0]
        if len(filename) == 0:
            return
        # открытие
        try:
            fin = open(filename, 'r')
        except Exception:
            QMessageBox.information(self, 'Ошибка', 'Ошибка открытия файла!')
            return
        # чтение, выполнение
        sql = QSqlQuery()

        cmd_arr = fin.read().split(';')
        fin.close()

        for cmd in cmd_arr:
            cmd = cmd.lstrip().rstrip()
            if len(cmd) == 0:
                continue
            cmd = cmd + ';'
            sql.exec(cmd)
            if sql.lastError().type() != QSqlError.NoError:
                QMessageBox.information(self, 'Ошибка', 'Ошибка выполнения запроса!\n'
                                                        'Ошибка: {0}\n'
                                                        'Запрос: {1}'.format(sql.lastError().text(), cmd))
                return
        QMessageBox.information(self, 'Информация', 'Импорт выполнен!')

        self.scheduleView.setModel(ScheduleModel())
        self.pulpitsView.setModel(PulpitsModel())
        self.groupsView.setModel(GroupsModel())

    # MENU EXCEL
    def slot_ExportExcel(self):
        if self.db == None:
            QMessageBox.information(self, 'Ошибка', 'Экспорт невозможен, т.к. БД не открыта!')
            return
        # получение имени файла
        dlg = QFileDialog()
        filename_cort = dlg.getSaveFileName(caption='Введите имя файла',
                                            filter='*.xlsx')
        filename = filename_cort[0]
        if len(filename) == 0:
            return
        # получение фильтра
        filter = self.filterDlg.getFilter()
        if filter == None:
            return
        # создание запроса
        template_query = 'SELECT schedule.id, pulpits.name, groups.name, ' \
                         'weeks.name, days.name, schedule.n_lesson, schedule.lesson ' \
                         'FROM schedule ' \
                         'INNER JOIN pulpits ON pulpits.id = schedule.id_pulpit ' \
                         'INNER JOIN groups ON groups.id = schedule.id_group ' \
                         'INNER JOIN weeks ON weeks.id = schedule.id_week ' \
                         'INNER JOIN days ON days.id = schedule.id_day '
        arr_fields = ['schedule.id', 'pulpits.name', 'groups.name',
                      'weeks.name', 'days.name', 'schedule.n_lesson',
                      'schedule.lesson']
        condition = 'WHERE '
        isFirst = True
        for i in range(7):
            if filter[i] == None:
                continue
            if not isFirst:
                condition = condition + ' AND '
            else:
                isFirst = False

            if (i >= 1 and i <= 4) or i == 6:
                condition = condition + arr_fields[i] + '=' + '"' + filter[i] + '"'
            else:
                condition = condition + arr_fields[i] + '=' + filter[i]

        query = ''
        if isFirst:
            query = template_query + ';'
        else:
            query = template_query + condition + ';'

        sql = QSqlQuery()
        sql.exec(query)
        if sql.lastError().type() != QSqlError.NoError:
            QMessageBox.information(self, 'Ошибка',
                                    'Ошибка выполнения запроса: {0}'
                                    .format(sql.lastError().text()))
            return
        # создание документа, запись строк и сохранение
        doc = Workbook()
        sheet = doc.active
        headers = ['id', 'Кафедра', 'Группа', 'Неделя', 'День', '№ пары', 'Пара']
        for i in range(1, 8):
            sheet.cell(1, i).set_explicit_value(headers[i - 1])

        row = 2
        while sql.next():
            for i in range(7):
                sheet.cell(row, i + 1).set_explicit_value(str(sql.value(i)))
            row += 1

        try:
            doc.save(filename)
        except Exception:
            QMessageBox.information(self, 'Ошибка', 'Ошибка сохранения документа!')

    def __insertRow(self, arr_values):
        '''Обрабатывает и добавляет массив значений в БД'''
        insert_schedule_tmp = 'INSERT INTO schedule (id, id_pulpit, id_group, id_week, id_day, n_lesson, lesson) ' \
                              'VALUES (?, ?, ?, ?, ?, ?, ?);'
        insert_pulpit_tmp = 'INSERT INTO pulpits (id, name) ' \
                            'VALUES (?, ?);'
        insert_group_tmp = 'INSERT INTO groups (id, name) ' \
                           'VALUES (?, ?);'
        translated_arr = [None, None, None, None, None, None]
        # 0-3 поиск(вставка в таб), 4 - проверка в range(1, 10), 5 - без проверки
        # обработка кафедры, группы, вставка если не найдены
        for i in range(2):
            id = self.__getIdSubtable(i + 1, arr_values[i])
            if id == None:
                sql = QSqlQuery()
                if i == 0:
                    sql.prepare(insert_pulpit_tmp)
                elif i == 1:
                    sql.prepare(insert_group_tmp)
                sql.bindValue(0, None)
                sql.bindValue(1, arr_values[i])
                sql.exec()
                if sql.lastError().type() != QSqlError.NoError:
                    return False
                translated_arr[i] = self.__getIdSubtable(i + 1, arr_values[i])
            else:
                translated_arr[i] = id
        # обработка недели, дня
        for i in range(2, 4):
            id = self.__getIdSubtable(i + 1, arr_values[i])
            if id == None:
                return False
            else:
                translated_arr[i] = id
        # обработка номера пары
        try:
            n_lesson = int(arr_values[4])
        except Exception:
            return False
        if n_lesson > 10 or n_lesson < 1:
            return False
        else:
            translated_arr[4] = n_lesson
        # название пары
        translated_arr[5] = arr_values[5]

        # вставка в БД
        sql = QSqlQuery()
        sql.prepare(insert_schedule_tmp)
        sql.bindValue(0, None)
        for i in range(6):
            sql.bindValue(i + 1, translated_arr[i])
        sql.exec()
        if sql.lastError().type() != QSqlError.NoError:
            return False
        return True

    def slot_ImportExcel(self):
        if self.db == None:
            QMessageBox.information(self, 'Ошибка', 'Импорт невозможен, т.к. БД не открыта!')
            return
        # получ имени файла
        dlg = QFileDialog()
        filename_cort = dlg.getOpenFileName(caption='Выберите файл',
                                            filter='*.xlsx')
        filename = filename_cort[0]
        if len(filename) == 0:
            return
        # загрузка документа
        try:
            doc = load_workbook(filename)
        except Exception:
            QMessageBox.information(self, 'Ошибка', 'Ошибка открытия документа!')
            return


        # сбор строк и вставка в БД
        for sheet in doc.worksheets:
            row = 2
            while row <= sheet.max_row:
                arr_values = []
                bad_cell = False
                for col in range(1, 7):
                    if sheet.cell(row, col + 1).value == None:
                        bad_cell = True
                        break
                    arr_values.append(str(sheet.cell(row, col + 1).value))
                row += 1
                # обрабатываю только строки со всеми ячейками
                if bad_cell:
                    continue
                if not self.__insertRow(arr_values):
                    QMessageBox.information(self, 'Ошибка', 'Ошибка вставки в БД!\n'
                                                            'Строка № {0}'.format(row - 1))
        doc.close()

        self.scheduleView.setModel(ScheduleModel())
        self.pulpitsView.setModel(PulpitsModel())
        self.groupsView.setModel(GroupsModel())


    def __getIdSubtable(self, colums: int, name: str):
        '''Запрос к одной из 4 таблиц, ищет строку в таблице, возвращает id/None'''
        if colums == None or name == None:
            return None
        # подготовка запроса
        template_query = 'SELECT id FROM {0} WHERE name="{1}";'
        if colums == 1:
            template_query = template_query.format('pulpits', name)
        elif colums == 2:
            template_query = template_query.format('groups', name)
        elif colums == 3:
            template_query = template_query.format('weeks', name)
        elif colums == 4:
            template_query = template_query.format('days', name)
        # выполнение запроса
        sql = QSqlQuery()
        sql.exec(template_query)
        if sql.lastError().type() != QSqlError.NoError or not sql.next():
            return None
        return sql.value(0)

    # MENU FILTER
    def slot_SetFilter(self):
        '''Выставляет фильтр для расписания'''
        filter = self.filterDlg.getFilter()
        if filter == None:
            return

        filter_str = ''
        isFirst = True
        arrFields = ['schedule.id', 'schedule.id_pulpit', 'schedule.id_group',
                     'schedule.id_week', 'schedule.id_day', 'schedule.n_lesson',
                     'schedule.lesson']
        try:
            for i in range(7):
                if filter[i] == None:
                    continue
                if not isFirst:
                    filter_str = filter_str + ' AND '
                else:
                    isFirst = False

                if i >= 1 and i <= 4:
                    id = self.__getIdSubtable(i, filter[i])
                    if id == None:
                        id = -1
                    filter_str = filter_str + arrFields[i] + '=' + str(id)
                elif i == 6:
                    filter_str = filter_str + arrFields[i] + '=' + '"' + str(filter[i]) + '"'
                else:
                    filter_str = filter_str + arrFields[i] + '=' + str(filter[i])
        except Exception:
            print(Exception)

        self.scheduleView.model().setFilter(filter_str)
        self.scheduleView.model().select()
        #print(filter_str)

    def slot_SetDefaultFilter(self):
        '''Отключение фильтра'''
        self.scheduleView.model().setFilter('')

    # MENU HELP
    def slot_Author(self):
        QMessageBox.information(self, 'Автор',
                                'Студент МАИ\n'
                                'Группа: М8О-213Б-17\n'
                                'ФИО: Семенов Сергей Дмитриевич')

    def slot_AboutProg(self):
        QMessageBox.information(self, 'Помощь',
                                'Программа открывает, редактирует БД с таблицами "Расписание", "Кафедры", "Группы".\n\n'
                                '   Кнопки:\n'
                                '1. "Записать" - применяет изменения в таблице к БД.\n'
                                '2. "Вернуть" - восстанавливает данные из БД.\n'
                                '3. "Новая строка" - добавляет строку в таблицу.\n'
                                '4. "Удалить строки" - удаляет строки из БД.\n\n'
                                '   Меню -> Файл:\n'
                                '1. "Создать пустую БД" - создает новую БД с указанным именем.\n'
                                '2. "Открыть БД" - открывает указанную БД.\n'
                                '3. "Сохранить БД" - сохраняет текущую БД в указанный файл.\n'
                                '4. "Создать таблицы" - создает в текущей БД все таблицы.\n'
                                '5. "Полный экспорт в *.sql" '
                                        '- экспорт команд создания таблиц и вставки данных в таблицы.\n'
                                '6. "Экспорт данных в *.sql" - экспорт команд вставки данных в таблицы.\n'
                                '7. "Импорт *.sql" - выполнение команд из файла к текущей БД.\n\n'
                                '   Меню -> Excel:\n'
                                '1. "Экспортировать по фильтру" '
                                        '- сохраняет строки из таблицы "Расписание" с указанным фильтром.\n'
                                '2. "Импортировать" - считывает данные из таблицы Excel и добавляет в таблицы БД. '
                                        'Если в строке встречена пустая ячейка, строка пропускается. '
                                        'Если кафедра или группа не найдены в соответствующих таблицах БД, то они добавляются в эти таблицы. '
                                        'Названия недель, дней должны точно совпадать с теми, что выбираются в ComboBox\'e.\n\n'
                                '   Меню -> Фильтр\n'
                                '1. "Выставить фильтр" - выставляет фильтр для таблицы "Расписание".\n'
                                '2. "Сбросить фильтр" - убирает фильтр.')



if __name__ == '__main__':
    import sys
    app = QApplication(sys.argv)
    wnd = SQLWindow()
    wnd.show()
    sys.exit(app.exec_())